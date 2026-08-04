"""
POST /api/v1/import/excel
Загрузка Excel → парсинг → валидация → сохранение в БД → расчёт аналитики
"""
import json
import io
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, delete
import openpyxl

from backend.core.database import get_db
from backend.models.models import (
    Import, Session, SessionTotals, SourceRow,
    SourceGroup, ValidationLog, AnalyticsCache, Manager
)
from backend.analytics.engine import run_import_pipeline
from backend.routers.analytics import compute_and_cache

router = APIRouter(tags=["Import"])

MONTH_MAP = {
    "январь": 1, "февраль": 2, "март": 3, "апрель": 4,
    "май": 5, "июнь": 6, "июль": 7, "август": 8,
    "сентябрь": 9, "октябрь": 10, "ноябрь": 11, "декабрь": 12,
}

DEPT_MANAGER_ID = uuid.UUID("00000000-0000-0000-0000-000000000001")


def parse_period(period: str) -> tuple[int, int]:
    """'Июнь 2026' → (2026, 6)"""
    parts = period.strip().split()
    if len(parts) != 2:
        raise ValueError(f"Неверный формат периода: {period}")
    month_name = parts[0].lower()
    year = int(parts[1])
    month = MONTH_MAP.get(month_name)
    if not month:
        raise ValueError(f"Неизвестный месяц: {parts[0]}")
    return year, month


@router.post("/import/excel")
async def import_excel(
    file:       UploadFile = File(...),
    manager_id: Optional[str] = Form(None),
    period:     str = Form(...),
    db:         AsyncSession = Depends(get_db),
):
    """
    Импорт Excel файла.
    - manager_id: UUID менеджера или null (весь отдел)
    - period: 'Июнь 2026'
    """
    # 1. Парсим период
    try:
        period_year, period_month = parse_period(period)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    # 2. Определяем менеджера
    mgr_id = DEPT_MANAGER_ID
    if manager_id and manager_id != "_dept":
        try:
            mgr_id = uuid.UUID(manager_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный manager_id")

    # 3. Читаем Excel
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else ""
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
        raw_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_rows.append(dict(zip(headers, row)))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    # Проверяем обязательные колонки
    required = ["Источник", "Лиды", "Продажи", "Выручка"]
    missing  = [c for c in required if c not in headers]
    if missing:
        raise HTTPException(status_code=400,
                            detail=f"Не хватает колонок: {', '.join(missing)}")

    # 4. Загружаем правила группировки из БД
    groups_result = await db.execute(
        select(SourceGroup).where(SourceGroup.is_active == True).order_by(SourceGroup.priority)
    )
    group_rules = [
        {
            "group_name": g.group_name,
            "match_type": g.match_type,
            "pattern":    g.pattern,
            "aliases":    g.aliases or [],
            "priority":   g.priority,
            "is_active":  g.is_active,
        }
        for g in groups_result.scalars().all()
    ]

    # 5. Pipeline: normalize → validate → enrich
    parsed_json = {"headers": headers, "rows": raw_rows}  # для диагностики
    result = run_import_pipeline(raw_rows, group_rules)

    quality = result["quality"]

    # 6. Сохраняем Import
    import_obj = Import(
        filename      = file.filename,
        manager_id    = mgr_id,
        period        = period,
        period_year   = period_year,
        period_month  = period_month,
        row_count     = quality["total"],
        valid_count   = quality["valid"],
        warn_count    = quality["warning"],
        invalid_count = quality["invalid"],
        status        = "processing",
        # raw_file_url и parsed_json_url заполним через Supabase Storage
    )
    db.add(import_obj)
    await db.flush()

    # 7. Upsert Session (менеджер + период уникальны)
    existing_sess = await db.execute(
        select(Session).where(
            Session.manager_id == mgr_id,
            Session.period == period
        )
    )
    sess = existing_sess.scalar_one_or_none()

    if sess:
        # Обновляем существующую сессию — удаляем старые данные
        sess.import_id = import_obj.id
        await db.execute(
            delete(SourceRow).where(SourceRow.session_id == sess.id)
        )
        await db.execute(
            delete(SessionTotals).where(SessionTotals.session_id == sess.id)
        )
        await db.execute(
            delete(ValidationLog).where(ValidationLog.import_id == import_obj.id)
        )
    else:
        sess = Session(
            import_id    = import_obj.id,
            manager_id   = mgr_id,
            period       = period,
            period_year  = period_year,
            period_month = period_month,
        )
        db.add(sess)
        await db.flush()

    # 8. Сохраняем SessionTotals (строка ИТОГО)
    t = result["totals"]
    totals_obj = SessionTotals(
        session_id    = sess.id,
        leads         = t.leads,
        new_leads     = t.new_leads,
        rep_leads     = t.rep_leads,
        deals         = t.deals,
        new_deals     = t.new_deals,
        rep_deals     = t.rep_deals,
        revenue       = t.revenue,
        rep_revenue   = t.rep_revenue,
        conv          = t.conv,
        avg_check     = t.avg_check,
        pct_rep_leads = t.pct_rep_leads,
        pct_rep_deals = t.pct_rep_deals,
        not_realized  = t.not_realized,
    )
    db.add(totals_obj)

    # 9. Сохраняем SourceRows
    all_rows = result["analytics_rows"] + result["invalid_rows"]
    row_objects = []
    for r in all_rows:
        row_obj = SourceRow(
            session_id      = sess.id,
            src             = r.src,
            campaign        = r.campaign,
            content         = r.content,
            group_name      = r.group_name,
            leads           = r.leads,
            new_leads       = r.new_leads,
            rep_leads       = r.rep_leads,
            deals           = r.deals,
            new_deals       = r.new_deals,
            rep_deals       = r.rep_deals,
            revenue         = r.revenue,
            rep_revenue     = r.rep_revenue,
            conv_excel      = r.conv_excel,
            conv_calc       = r.conv_calc,
            conv            = r.conv,
            avg_check_excel = r.avg_check_excel,
            avg_check       = r.avg_check,
            pct_rep_deals   = r.pct_rep_deals,
            pct_rep_leads   = r.pct_rep_leads,
            not_realized    = r.not_realized,
            status          = r.status,
            issues          = r.issues,
            is_manual       = False,
        )
        db.add(row_obj)
        row_objects.append(row_obj)

    # 10. Сохраняем ValidationLog (все ошибки)
    for r in all_rows:
        for issue in r.issues:
            vlog = ValidationLog(
                import_id  = import_obj.id,
                row_number = r.row_number,
                src        = r.src,
                rule_code  = issue["code"],
                severity   = issue["severity"],
                message    = issue["msg"],
                raw_data   = r.raw_data,
            )
            db.add(vlog)

    await db.flush()

    # 11. Финальная аналитика → analytics_cache
    import_obj.status = "done"
    await db.commit()

    # Пересчитываем кэш (отдельный запрос после commit)
    await compute_and_cache(str(sess.id), db)

    return {
        "session_id":  str(sess.id),
        "import_id":   str(import_obj.id),
        "period":      period,
        "manager_id":  str(mgr_id),
        "quality":     quality,
        "has_itogo":   result["has_itogo"],
        "message":     f"Импортировано: {quality['total']} строк. "
                       f"VALID: {quality['valid']}, "
                       f"WARNING: {quality['warning']}, "
                       f"INVALID: {quality['invalid']}",
    }


# ============================================================
# POST /api/v1/import/tree-json
# Прямой импорт живого отчёта CRM (report/tree), без async-экспорта
# и без Excel. Устраняет расхождение между асинхронной выгрузкой и
# тем, что реально показывает экран "Сквозная аналитика" в CRM
# (баг подтверждён: async export завышал выручку на 10-70%).
# ============================================================
from pydantic import BaseModel, Field, ValidationError
from typing import Any, Dict, List, Optional


class TreeTotalIn(BaseModel):
    """
    Строка "Итого" из ответа CRM report/tree?view=utm.
    leads/sales/revenue обязательны (без дефолта) — это три ключевых KPI,
    которые пишутся в session_totals и напрямую управляют всем сайтом. Если
    CRM когда-нибудь переименует одно из этих полей в своём API, раньше
    _num(total.get(...)) молча превращал бы отсутствующее поле в 0 для ВСЕХ
    менеджеров без единой ошибки — теперь Pydantic вернёт понятную 422 ошибку
    с именем недостающего поля вместо тихой порчи данных.
    """
    leads: float
    sales: float
    revenue: float
    new_leads: float = 0
    repeat_leads: float = 0
    new_sales: float = 0
    repeat_sales: float = 0
    repeat_revenue: float = 0
    conversion: float = 0
    average_check: float = 0
    procent_of_repeat_leads: float = 0
    procent_of_repeat_sales: float = 0
    procent_of_new_sales: float = 0
    not_implemented: float = 0


class TreeRowMetricsIn(BaseModel):
    """То же самое на уровне отдельного источника (data[].metrics) — здесь
    все поля с дефолтом 0, потому что отдельный источник с нулями — обычное,
    легитимное состояние (см. аудит), в отличие от отсутствующего total."""
    leads: float = 0
    new_leads: float = 0
    repeat_leads: float = 0
    sales: float = 0
    new_sales: float = 0
    repeat_sales: float = 0
    revenue: float = 0
    repeat_revenue: float = 0
    conversion: float = 0
    average_check: float = 0
    procent_of_repeat_leads: float = 0
    procent_of_repeat_sales: float = 0
    not_implemented: float = 0


class TreeRowIn(BaseModel):
    label: Optional[str] = None
    metrics: TreeRowMetricsIn = Field(default_factory=TreeRowMetricsIn)


class TreeDataIn(BaseModel):
    total: TreeTotalIn
    data: List[TreeRowIn] = Field(default_factory=list)


class TreeImportIn(BaseModel):
    manager_id: str
    period: str
    tree: TreeDataIn


@router.post("/import/tree-json")
async def import_tree_json(
    body: TreeImportIn,
    db: AsyncSession = Depends(get_db),
):
    """
    body.tree — сырой ответ CRM-эндпоинта
    /api/analytics/report/tree?view=utm&pineline_id=...&manager_id=...
    (см. поля total{} и data[]). Собирается напрямую в браузере
    (live-запрос, тот же, что рисует экран), без промежуточного
    xlsx-файла и без /export/store — единственный источник для Сквозной
    аналитики с этого момента.
    """
    try:
        period_year, period_month = parse_period(body.period)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    mgr_id = DEPT_MANAGER_ID
    if body.manager_id and body.manager_id != "_dept":
        try:
            mgr_id = uuid.UUID(body.manager_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Неверный manager_id")

    total = body.tree.total.model_dump()
    data_rows: List[Dict[str, Any]] = [
        {"label": r.label, "metrics": r.metrics.model_dump()} for r in body.tree.data
    ]

    # Upsert Session
    existing_sess = await db.execute(
        select(Session).where(
            Session.manager_id == mgr_id,
            Session.period == body.period
        )
    )
    sess = existing_sess.scalar_one_or_none()

    if sess:
        await db.execute(delete(SourceRow).where(SourceRow.session_id == sess.id))
        await db.execute(delete(SessionTotals).where(SessionTotals.session_id == sess.id))
    else:
        sess = Session(
            import_id    = None,
            manager_id   = mgr_id,
            period       = body.period,
            period_year  = period_year,
            period_month = period_month,
        )
        db.add(sess)
        await db.flush()

    def _num(v):
        return v if v is not None else 0

    # conv = поле CRM "conversion" как есть — CRM уже считает его как sales/leads
    # за период по date_filter=created_at (та же когорта "лидов, поступивших в
    # этот период"), это и есть готовая конверсия, показанная в самой CRM.
    totals_obj = SessionTotals(
        session_id    = sess.id,
        leads         = _num(total.get("leads")),
        new_leads     = _num(total.get("new_leads")),
        rep_leads     = _num(total.get("repeat_leads")),
        deals         = _num(total.get("sales")),
        new_deals     = _num(total.get("new_sales")),
        rep_deals     = _num(total.get("repeat_sales")),
        revenue       = _num(total.get("revenue")),
        rep_revenue   = _num(total.get("repeat_revenue")),
        conv          = _num(total.get("conversion")),
        avg_check     = _num(total.get("average_check")),
        pct_rep_leads = _num(total.get("procent_of_repeat_leads")),
        pct_rep_deals = _num(total.get("procent_of_repeat_sales")),
        not_realized  = _num(total.get("not_implemented")),
    )
    db.add(totals_obj)

    row_count = 0
    for item in data_rows:
        m = item.get("metrics") or {}
        label = item.get("label") or "Без источника"
        row_obj = SourceRow(
            session_id      = sess.id,
            src             = str(label),
            campaign        = None,
            content         = None,
            group_name      = None,
            leads           = _num(m.get("leads")),
            new_leads       = _num(m.get("new_leads")),
            rep_leads       = _num(m.get("repeat_leads")),
            deals           = _num(m.get("sales")),
            new_deals       = _num(m.get("new_sales")),
            rep_deals       = _num(m.get("repeat_sales")),
            revenue         = _num(m.get("revenue")),
            rep_revenue     = _num(m.get("repeat_revenue")),
            conv_excel      = _num(m.get("conversion")),
            conv_calc       = _num(m.get("conversion")),
            conv            = _num(m.get("conversion")),
            avg_check_excel = _num(m.get("average_check")),
            avg_check       = _num(m.get("average_check")),
            pct_rep_deals   = _num(m.get("procent_of_repeat_sales")),
            pct_rep_leads   = _num(m.get("procent_of_repeat_leads")),
            not_realized    = _num(m.get("not_implemented")),
            status          = "VALID",
            issues          = [],
            is_manual       = False,
        )
        db.add(row_obj)
        row_count += 1

    await db.commit()
    await compute_and_cache(str(sess.id), db)

    return {
        "session_id": str(sess.id),
        "period": body.period,
        "manager_id": str(mgr_id),
        "source": "live_tree",
        "rows_imported": row_count,
        "totals": {
            "leads": total.get("leads"),
            "sales": total.get("sales"),
            "revenue": total.get("revenue"),
            "conversion_new": total.get("procent_of_new_sales"),
        },
        "message": f"Импортировано из живого отчёта (report/tree): {row_count} источников, "
                   f"лидов {total.get('leads')}, сделок {total.get('sales')}, "
                   f"выручка {total.get('revenue')} ₸",
    }
