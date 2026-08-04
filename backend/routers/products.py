"""
Товарная аналитика по менеджерам (категории/конкретные товары),
собранная из CRM Zymyran (Аналитика -> Товары, разрез по дате доставки).

POST   /api/v1/import/products         — загрузка Excel (3 листа: По месяцам / По категориям / По товарам)
GET    /api/v1/products/{manager_id}   — данные для карточки менеджера (с применёнными правками)
POST   /api/v1/products/override       — точечная правка значения (переживает повторный импорт)
DELETE /api/v1/products/override/{id}  — снять правку, вернуться к импортированному значению
"""
import io
import re
import uuid
from collections import defaultdict
from typing import Optional

from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import openpyxl

from backend.core.database import get_db
from backend.models.models import (
    Manager, ProductImport, ProductSalesSummary,
    ProductSalesCategory, ProductSalesItem, ProductOverride,
)
from backend.routers.imports import MONTH_MAP

router = APIRouter(tags=["Products"])

REQUIRED_SHEETS = ["По месяцам", "По категориям", "По товарам"]
NUMERIC_FIELDS = {"positions", "qty", "revenue"}


def _norm_name(name: str) -> str:
    """Нормализация имени менеджера/категории/товара для сопоставления: убираем лишние пробелы, приводим к нижнему регистру."""
    return re.sub(r"\s+", " ", (name or "").strip()).lower()


def _sheet_rows(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if all(v is None for v in d.values()):
            continue
        rows.append(d)
    return rows


def _num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0


@router.post("/import/products")
async def import_products(
    file: UploadFile = File(...),
    year: int = Form(...),
    db: AsyncSession = Depends(get_db),
):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Ошибка чтения Excel: {e}")

    missing_sheets = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing_sheets:
        raise HTTPException(status_code=400, detail=f"Не хватает листов: {', '.join(missing_sheets)}")

    # --- Менеджеры из БД, нормализованные для сопоставления по имени ---
    mgr_result = await db.execute(select(Manager))
    managers = mgr_result.scalars().all()
    name_to_id = {_norm_name(m.name): m.id for m in managers}

    unmatched_names = set()

    def resolve_manager(raw_name: str):
        mid = name_to_id.get(_norm_name(raw_name))
        if mid is None:
            unmatched_names.add(raw_name)
        return mid

    def resolve_period(month_name: str):
        key = _norm_name(month_name)
        month_num = MONTH_MAP.get(key)
        if not month_num:
            return None, None
        period_str = f"{month_name.strip().capitalize()} {year}"
        return period_str, month_num

    # --- Лист "По месяцам" -> сводка по (manager_id, period) ---
    summaries = {}  # (manager_id, period) -> dict
    for row in _sheet_rows(wb["По месяцам"]):
        mid = resolve_manager(row.get("Менеджер", ""))
        period, month_num = resolve_period(row.get("Месяц", ""))
        if mid is None or period is None:
            continue
        summaries[(mid, period)] = {
            "period_year": year,
            "period_month": month_num,
            "positions": _num(row.get("Позиций")),
            "qty": _num(row.get("Кол-во товара, шт")),
            "revenue": _num(row.get("Выручка, ₸")),
        }

    # --- Лист "По категориям" -> список категорий по (manager_id, period), в порядке файла ---
    categories = defaultdict(list)
    for row in _sheet_rows(wb["По категориям"]):
        mid = resolve_manager(row.get("Менеджер", ""))
        period, _ = resolve_period(row.get("Месяц", ""))
        if mid is None or period is None:
            continue
        categories[(mid, period)].append({
            "category": str(row.get("Категория товара") or "").strip(),
            "qty": _num(row.get("Кол-во, шт")),
            "revenue": _num(row.get("Выручка, ₸")),
        })

    # --- Лист "По товарам" -> список товаров по (manager_id, period), в порядке файла ---
    items = defaultdict(list)
    for row in _sheet_rows(wb["По товарам"]):
        mid = resolve_manager(row.get("Менеджер", ""))
        period, _ = resolve_period(row.get("Месяц", ""))
        if mid is None or period is None:
            continue
        items[(mid, period)].append({
            "product_name": str(row.get("Товар") or "").strip(),
            "qty": _num(row.get("Кол-во, шт")),
            "revenue": _num(row.get("Выручка, ₸")),
        })

    if not summaries:
        raise HTTPException(
            status_code=400,
            detail="Не удалось сопоставить ни одной строки. Проверьте имена менеджеров: "
                   + ", ".join(sorted(unmatched_names)) if unmatched_names else
                   "Лист «По месяцам» пуст или колонки названы иначе.",
        )

    # --- Import batch ---
    imp = ProductImport(
        filename=file.filename,
        manager_count=len({mid for (mid, _period) in summaries.keys()}),
        row_count=len(summaries) + sum(len(v) for v in categories.values()) + sum(len(v) for v in items.values()),
        status="done",
    )
    db.add(imp)
    await db.flush()

    # --- Upsert по (manager_id, period); категории/товары апсертятся ПО ИМЕНИ, чтобы их id
    #     оставался стабильным между загрузками — иначе точечные правки (product_overrides)
    #     будут "отвязываться" от строки при каждом повторном импорте (CRM-данные могут
    #     обновляться задним числом даже за прошлые месяцы). Строки, у которых есть активная
    #     правка, но которых больше нет в свежей выгрузке, не удаляются — чтобы не терять правку.
    for (mid, period), vals in summaries.items():
        existing = await db.execute(
            select(ProductSalesSummary).where(
                ProductSalesSummary.manager_id == mid,
                ProductSalesSummary.period == period,
            )
        )
        summary_obj = existing.scalar_one_or_none()
        if summary_obj:
            summary_obj.import_id = imp.id
            summary_obj.positions = vals["positions"]
            summary_obj.qty = vals["qty"]
            summary_obj.revenue = vals["revenue"]
            summary_obj.period_year = vals["period_year"]
            summary_obj.period_month = vals["period_month"]
        else:
            summary_obj = ProductSalesSummary(
                import_id=imp.id,
                manager_id=mid,
                period=period,
                period_year=vals["period_year"],
                period_month=vals["period_month"],
                positions=vals["positions"],
                qty=vals["qty"],
                revenue=vals["revenue"],
            )
            db.add(summary_obj)
            await db.flush()

        # -- категории: апсерт по имени --
        cat_result = await db.execute(
            select(ProductSalesCategory).where(ProductSalesCategory.summary_id == summary_obj.id)
        )
        existing_cats = {c.category: c for c in cat_result.scalars().all()}
        seen_cat_names = set()
        for rank, cat in enumerate(categories.get((mid, period), []), start=1):
            seen_cat_names.add(cat["category"])
            row_obj = existing_cats.get(cat["category"])
            if row_obj:
                row_obj.qty = cat["qty"]
                row_obj.revenue = cat["revenue"]
                row_obj.rank = rank
            else:
                db.add(ProductSalesCategory(
                    summary_id=summary_obj.id,
                    category=cat["category"],
                    qty=cat["qty"],
                    revenue=cat["revenue"],
                    rank=rank,
                ))
        for name, row_obj in existing_cats.items():
            if name in seen_cat_names:
                continue
            has_override = await db.execute(
                select(ProductOverride).where(
                    ProductOverride.target_type == "category",
                    ProductOverride.target_id == row_obj.id,
                )
            )
            if not has_override.scalars().first():
                await db.delete(row_obj)

        # -- товары: апсерт по имени --
        item_result = await db.execute(
            select(ProductSalesItem).where(ProductSalesItem.summary_id == summary_obj.id)
        )
        existing_items = {i.product_name: i for i in item_result.scalars().all()}
        seen_item_names = set()
        for rank, it in enumerate(items.get((mid, period), []), start=1):
            seen_item_names.add(it["product_name"])
            row_obj = existing_items.get(it["product_name"])
            if row_obj:
                row_obj.qty = it["qty"]
                row_obj.revenue = it["revenue"]
                row_obj.rank = rank
            else:
                db.add(ProductSalesItem(
                    summary_id=summary_obj.id,
                    product_name=it["product_name"],
                    qty=it["qty"],
                    revenue=it["revenue"],
                    rank=rank,
                ))
        for name, row_obj in existing_items.items():
            if name in seen_item_names:
                continue
            has_override = await db.execute(
                select(ProductOverride).where(
                    ProductOverride.target_type == "item",
                    ProductOverride.target_id == row_obj.id,
                )
            )
            if not has_override.scalars().first():
                await db.delete(row_obj)

    await db.commit()

    matched_names = sorted({m.name for m in managers if m.id in {mid for (mid, _p) in summaries.keys()}})

    return {
        "import_id": str(imp.id),
        "periods": sorted({p for (_m, p) in summaries.keys()}),
        "managers_matched": matched_names,
        "managers_unmatched": sorted(unmatched_names),
        "summary_count": len(summaries),
        "category_count": sum(len(v) for v in categories.values()),
        "item_count": sum(len(v) for v in items.values()),
        "message": f"Импортировано: {len(summaries)} пар менеджер/период."
                   + (f" Не сопоставлены: {', '.join(sorted(unmatched_names))}." if unmatched_names else ""),
    }


async def _overrides_for(db: AsyncSession, target_type: str, target_ids):
    if not target_ids:
        return {}
    result = await db.execute(
        select(ProductOverride).where(
            ProductOverride.target_type == target_type,
            ProductOverride.target_id.in_(target_ids),
        )
    )
    out = defaultdict(dict)
    for ov in result.scalars().all():
        out[ov.target_id][ov.field] = ov
    return out


def _apply_override(raw_value, field: str, override_obj):
    if override_obj is None:
        return raw_value, False
    if field in NUMERIC_FIELDS:
        try:
            return float(override_obj.value), True
        except (TypeError, ValueError):
            return raw_value, False
    return override_obj.value, True


@router.get("/products/{manager_id}")
async def get_manager_products(
    manager_id: str,
    period: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    try:
        mid = uuid.UUID(manager_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный manager_id")

    q = select(ProductSalesSummary).where(ProductSalesSummary.manager_id == mid)
    if period:
        q = q.where(ProductSalesSummary.period == period)
    result = await db.execute(q)
    summary_rows = result.scalars().all()

    summary_overrides = await _overrides_for(db, "summary", [s.id for s in summary_rows])

    periods_out = []
    for s in summary_rows:
        cat_result = await db.execute(
            select(ProductSalesCategory)
            .where(ProductSalesCategory.summary_id == s.id)
            .order_by(ProductSalesCategory.rank)
        )
        cat_rows = cat_result.scalars().all()
        item_result = await db.execute(
            select(ProductSalesItem)
            .where(ProductSalesItem.summary_id == s.id)
            .order_by(ProductSalesItem.rank)
        )
        item_rows = item_result.scalars().all()

        cat_overrides = await _overrides_for(db, "category", [c.id for c in cat_rows])
        item_overrides = await _overrides_for(db, "item", [i.id for i in item_rows])

        s_ov = summary_overrides.get(s.id, {})
        positions, pos_edited = _apply_override(float(s.positions or 0), "positions", s_ov.get("positions"))
        qty, qty_edited = _apply_override(float(s.qty or 0), "qty", s_ov.get("qty"))
        revenue, rev_edited = _apply_override(float(s.revenue or 0), "revenue", s_ov.get("revenue"))

        def cat_to_json(c):
            ov = cat_overrides.get(c.id, {})
            qty_v, qty_e = _apply_override(float(c.qty or 0), "qty", ov.get("qty"))
            rev_v, rev_e = _apply_override(float(c.revenue or 0), "revenue", ov.get("revenue"))
            name_v, name_e = _apply_override(c.category, "category", ov.get("category"))
            return {
                "id": str(c.id), "category": name_v, "qty": qty_v, "revenue": rev_v,
                "edited": bool(qty_e or rev_e or name_e),
            }

        def item_to_json(i):
            ov = item_overrides.get(i.id, {})
            qty_v, qty_e = _apply_override(float(i.qty or 0), "qty", ov.get("qty"))
            rev_v, rev_e = _apply_override(float(i.revenue or 0), "revenue", ov.get("revenue"))
            name_v, name_e = _apply_override(i.product_name, "product_name", ov.get("product_name"))
            return {
                "id": str(i.id), "product_name": name_v, "qty": qty_v, "revenue": rev_v,
                "edited": bool(qty_e or rev_e or name_e),
            }

        periods_out.append({
            "summary_id": str(s.id),
            "period": s.period,
            "period_year": s.period_year,
            "period_month": s.period_month,
            "positions": positions,
            "qty": qty,
            "revenue": revenue,
            "edited": bool(pos_edited or qty_edited or rev_edited),
            "categories": [cat_to_json(c) for c in cat_rows],
            "items": [item_to_json(i) for i in item_rows],
        })

    periods_out.sort(key=lambda p: (p["period_year"], p["period_month"]))

    return {"manager_id": manager_id, "periods": periods_out}


class OverrideIn(BaseModel):
    target_type: str   # 'summary' | 'category' | 'item'
    target_id: str
    field: str
    value: str


@router.post("/products/override")
async def set_product_override(data: OverrideIn, db: AsyncSession = Depends(get_db)):
    if data.target_type not in ("summary", "category", "item"):
        raise HTTPException(status_code=400, detail="target_type должен быть summary/category/item")
    try:
        target_uuid = uuid.UUID(data.target_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный target_id")

    existing = await db.execute(
        select(ProductOverride).where(
            ProductOverride.target_type == data.target_type,
            ProductOverride.target_id == target_uuid,
            ProductOverride.field == data.field,
        )
    )
    ov = existing.scalar_one_or_none()
    if ov:
        ov.value = data.value
    else:
        ov = ProductOverride(
            target_type=data.target_type,
            target_id=target_uuid,
            field=data.field,
            value=data.value,
        )
        db.add(ov)
    await db.commit()
    return {"id": str(ov.id), "target_type": ov.target_type, "target_id": str(ov.target_id), "field": ov.field, "value": ov.value}


@router.delete("/products/override")
async def delete_product_override(target_type: str, target_id: str, field: str, db: AsyncSession = Depends(get_db)):
    try:
        target_uuid = uuid.UUID(target_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный target_id")
    existing = await db.execute(
        select(ProductOverride).where(
            ProductOverride.target_type == target_type,
            ProductOverride.target_id == target_uuid,
            ProductOverride.field == field,
        )
    )
    ov = existing.scalar_one_or_none()
    if not ov:
        raise HTTPException(status_code=404, detail="Правка не найдена")
    await db.delete(ov)
    await db.commit()
    return {"deleted": True}
